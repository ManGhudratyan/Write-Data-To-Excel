import openpyxl
from openpyxl.styles import Font, PatternFill

def write_data_to_excel_file(data, filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    header = ['Name', 'Surname', 'Age', 'Profession']
    bold_font = Font(bold=True)
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00')
    green_fill = PatternFill(start_color='006400', end_color='006400')

    sheet.append(header)
    for cell in sheet[1]:
        cell.font = bold_font
        cell.fill = yellow_fill

    for row in data:
        row_values = [row['name'], row['surname'], row['age'], row['profession']]
        sheet.append(row_values)

        if row['age'] > 25:
            for cell in sheet[sheet.max_row]:
                cell.fill = green_fill

    workbook.save(filename)
    workbook.close()

def main():
    data = [
        {'name': 'Mane', 'surname': 'Ghazaryan', 'age': 20, 'profession': 'programmer'},
        {'name': 'Arpi', 'surname': 'Isahakyan', 'age': 32, 'profession': 'singer'},
        {'name': 'Sargis', 'surname': 'Manukyan', 'age': 25, 'profession': 'teacher'},
        {'name': 'Tigran', 'surname': 'Hakobyan', 'age': 23, 'profession': 'barber'},
        {'name': 'Hovhannes', 'surname': 'Barseghyan', 'age': 48, 'profession': 'driver'},
        {'name': 'Van', 'surname': 'Sahakyan', 'age': 26, 'profession': 'writer'}
    ]

    user_choice = input('Sort by name, surname, age, or profession?: ')

    if user_choice in ['name', 'surname', 'age', 'profession']:
        new_file_name = f"sorted_by_{user_choice}.xlsx"
        sorted_data = sorted(data, key=lambda x: x[user_choice])
        write_data_to_excel_file(sorted_data, new_file_name)
    else:
        print('Please write only name, surname, age, or profession.')

if __name__ == "__main__":
    main()
